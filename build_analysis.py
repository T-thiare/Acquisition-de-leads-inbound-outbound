import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
import warnings
warnings.filterwarnings("ignore")

# ── LOAD ────────────────────────────────────────────────────────────────────
df = pd.read_csv("/home/claude/vroom-portfolio/data/leads_raw.csv", parse_dates=["date"])

# ── DERIVED METRICS ─────────────────────────────────────────────────────────
df["qualified"]  = df["status"].isin(["RDV planifié","RDV effectué","Signé"])
df["closed"]     = df["status"] == "Signé"
df["rdv"]        = df["status"].isin(["RDV planifié","RDV effectué","Signé"])
df["in_out"]     = df["inbound"].map({True:"Inbound",False:"Outbound"})

SOURCE_ORDER = ["LeBonCoin","LaCentrale","GBP","Facebook","Formulaire","Partenariat"]
MONTH_LABELS = ["Jan","Fév","Mar","Avr","Mai","Jun","Jul","Aoû","Sep","Oct","Nov","Déc"]

# ── COLOURS ─────────────────────────────────────────────────────────────────
C_NAVY    = "11182C"
C_GOLD    = "E0BC00"
C_NAVY_L  = "1A2440"
C_GREEN   = "10B981"
C_MUTED   = "6B7280"
C_BG      = "F3F4F6"
C_WHITE   = "FFFFFF"
C_RED     = "EF4444"
C_AMBER   = "F59E0B"
C_INBOUND = "1D9E75"
C_OUTBOUND= "94A3B8"
C_GOLD_L  = "FFF9C4"

def hdr(c="FFFFFF", bg=C_NAVY, bold=True, sz=10, wrap=False, halign="center"):
    f = Font(name="Arial", bold=bold, color=c, size=sz)
    fi = PatternFill("solid", fgColor=bg)
    a = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    return f, fi, a

def cell_style(ws, cell_ref, value=None, font_c="000000", bg=None,
               bold=False, sz=10, halign="center", valign="center",
               num_fmt=None, wrap=False, border=None):
    c = ws[cell_ref]
    if value is not None: c.value = value
    c.font = Font(name="Arial", color=font_c, bold=bold, size=sz)
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if num_fmt: c.number_format = num_fmt
    return c

def thin_border():
    s = Side(style="thin", color="E5E7EB")
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border(color="E5E7EB"):
    s = Side(style="thin", color=color)
    return Border(bottom=s)

def apply_table(ws, rows_range, cols, alt_bg="F9FAFB"):
    for r in rows_range:
        for c in cols:
            cell = ws.cell(row=r, column=c)
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=alt_bg)
            cell.border = thin_border()

# ══════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ── SHEET 1: RAW DATA (sample) ───────────────────────────────────────────────
ws_raw = wb.active
ws_raw.title = "📋 Données brutes"
ws_raw.sheet_view.showGridLines = False
ws_raw.freeze_panes = "A3"

cols_raw = ["lead_id","date","month","source","inbound","city","brand",
            "status","delay_days","cpl_eur","vehicle_value","margin_eur","margin_pct"]
headers_raw = ["ID Lead","Date","Mois","Source","Inbound","Ville","Marque",
               "Statut","Délai (j)","CPL (€)","Valeur véhicule","Marge (€)","Marge %"]

# Title row
ws_raw.merge_cells("A1:M1")
c = ws_raw["A1"]
c.value = "VROOM MARKET — Dataset Leads 2025  |  1 019 leads simulés · Données réalistes basées sur benchmarks marché"
c.font = Font(name="Arial", bold=True, size=11, color=C_WHITE)
c.fill = PatternFill("solid", fgColor=C_NAVY)
c.alignment = Alignment(horizontal="left", vertical="center")
ws_raw.row_dimensions[1].height = 28

# Headers row 2
for i, h in enumerate(headers_raw, 1):
    c = ws_raw.cell(row=2, column=i, value=h)
    c.font = Font(name="Arial", bold=True, size=9, color=C_WHITE)
    c.fill = PatternFill("solid", fgColor=C_NAVY_L)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
ws_raw.row_dimensions[2].height = 22

STATUS_COLORS = {
    "NRP":          "FEE2E2",
    "En cours":     "FEF3C7",
    "RDV planifié": "DBEAFE",
    "RDV effectué": "E0E7FF",
    "Signé":        "D1FAE5",
    "Perdu":        "F3F4F6",
}
SOURCE_COLORS = {
    "LeBonCoin":  "FFF9C4",
    "LaCentrale": "FFF3E0",
    "GBP":        "D1FAE5",
    "Facebook":   "EDE9FE",
    "Formulaire": "DBEAFE",
    "Partenariat":"ECFDF5",
}

sample = df.sample(300, random_state=42).sort_values("date").reset_index(drop=True)
for r, row in sample.iterrows():
    excel_r = r + 3
    vals = [row[c] for c in cols_raw]
    for ci, v in enumerate(vals, 1):
        cell = ws_raw.cell(row=excel_r, column=ci)
        if pd.isna(v): cell.value = ""
        elif cols_raw[ci-1] == "date": cell.value = v.strftime("%d/%m/%Y")
        elif cols_raw[ci-1] == "inbound": cell.value = "✓" if v else "—"
        elif cols_raw[ci-1] == "margin_pct" and v: cell.value = v; cell.number_format = "0.0%"
        elif cols_raw[ci-1] in ("vehicle_value","margin_eur","cpl_eur") and v:
            cell.value = v; cell.number_format = "#,##0 €"
        else: cell.value = v

        cell.font = Font(name="Arial", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()

        if cols_raw[ci-1] == "status" and str(v) in STATUS_COLORS:
            cell.fill = PatternFill("solid", fgColor=STATUS_COLORS[str(v)])
        elif cols_raw[ci-1] == "source" and str(v) in SOURCE_COLORS:
            cell.fill = PatternFill("solid", fgColor=SOURCE_COLORS[str(v)])
        elif excel_r % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="F9FAFB")

    ws_raw.row_dimensions[excel_r].height = 17

widths_raw = [10,11,8,12,8,13,12,14,9,8,14,10,8]
for i, w in enumerate(widths_raw, 1):
    ws_raw.column_dimensions[get_column_letter(i)].width = w

# ── SHEET 2: KPI DASHBOARD ───────────────────────────────────────────────────
ws_kpi = wb.create_sheet("📊 KPI Dashboard")
ws_kpi.sheet_view.showGridLines = False

# Title
ws_kpi.merge_cells("A1:N1")
c = ws_kpi["A1"]
c.value = "VROOM MARKET  ·  Tableau de bord Performance Leads  ·  Janvier – Décembre 2025"
c.font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
c.fill = PatternFill("solid", fgColor=C_NAVY)
c.alignment = Alignment(horizontal="left", vertical="center")
ws_kpi.row_dimensions[1].height = 34

# ── KPI CARDS ROW ──────────────────────────────────────────────────────────
total_leads = len(df)
total_signed = df["closed"].sum()
total_inbound = df["inbound"].sum()
pct_inbound = total_inbound / total_leads
total_rdv = df["rdv"].sum()
avg_delay = df["delay_days"].mean()
total_margin = df["margin_eur"].sum()
avg_cpl_out = df[~df["inbound"]]["cpl_eur"].mean()
conv_rate = total_signed / total_leads
inbound_conv = df[df["inbound"]]["closed"].mean()
outbound_conv = df[~df["inbound"]]["closed"].mean()

kpi_cards = [
    ("Total Leads", total_leads, "#,##0", C_NAVY, C_WHITE),
    ("Deals Signés", total_signed, "#,##0", C_GREEN, C_WHITE),
    ("Taux Closing Global", conv_rate, "0.0%", C_NAVY_L, C_WHITE),
    ("Conv. Inbound", inbound_conv, "0.0%", C_INBOUND, C_WHITE),
    ("Conv. Outbound", outbound_conv, "0.0%", C_OUTBOUND, "000000"),
    ("Part Inbound", pct_inbound, "0.0%", C_GOLD, C_NAVY),
    ("Délai Moyen (j)", avg_delay, "0.0", C_AMBER, C_WHITE),
    ("Marge Totale", total_margin, "#,##0 €", "065F46", C_WHITE),
]

ws_kpi.row_dimensions[3].height = 18
ws_kpi.row_dimensions[4].height = 40
ws_kpi.row_dimensions[5].height = 22
ws_kpi.row_dimensions[6].height = 14

col = 1
for label, val, fmt, bg, fc in kpi_cards:
    cl = get_column_letter(col)
    cl2 = get_column_letter(col+1)
    ws_kpi.merge_cells(f"{cl}3:{cl2}3")
    ws_kpi.merge_cells(f"{cl}4:{cl2}4")
    ws_kpi.merge_cells(f"{cl}5:{cl2}5")

    # Label
    lc = ws_kpi[f"{cl}3"]
    lc.value = label
    lc.font = Font(name="Arial", size=8, bold=True, color="9CA3AF")
    lc.fill = PatternFill("solid", fgColor="F9FAFB")
    lc.alignment = Alignment(horizontal="center", vertical="center")

    # Value
    vc = ws_kpi[f"{cl}4"]
    vc.value = val
    vc.number_format = fmt
    vc.font = Font(name="Arial", size=18, bold=True, color=fc)
    vc.fill = PatternFill("solid", fgColor=bg)
    vc.alignment = Alignment(horizontal="center", vertical="center")

    # Bottom accent bar
    bc = ws_kpi[f"{cl}5"]
    bc.fill = PatternFill("solid", fgColor=C_GOLD if bg != C_GOLD else C_NAVY)

    col += 2

ws_kpi.row_dimensions[7].height = 16

# ── MONTHLY TABLE ──────────────────────────────────────────────────────────
monthly = df.groupby("month_num").agg(
    total_leads=("lead_id","count"),
    inbound_leads=("inbound","sum"),
    signed=("closed","sum"),
    rdv=("rdv","sum"),
    avg_delay=("delay_days","mean"),
    margin=("margin_eur","sum"),
).reset_index()
monthly["outbound_leads"] = monthly["total_leads"] - monthly["inbound_leads"]
monthly["pct_inbound"] = monthly["inbound_leads"] / monthly["total_leads"]
monthly["conv_rate"] = monthly["signed"] / monthly["total_leads"]
monthly["month_label"] = [MONTH_LABELS[m-1] for m in monthly["month_num"]]

# Headers
r = 8
ws_kpi.merge_cells(f"A{r}:N{r}")
c = ws_kpi[f"A{r}"]
c.value = "ÉVOLUTION MENSUELLE DES LEADS — 2025"
c.font = Font(name="Arial", bold=True, size=10, color=C_WHITE)
c.fill = PatternFill("solid", fgColor=C_NAVY_L)
c.alignment = Alignment(horizontal="left", vertical="center")
ws_kpi.row_dimensions[r].height = 22

r = 9
monthly_hdrs = ["Mois","Total Leads","Outbound","Inbound","% Inbound",
                "RDV","Signés","Tx Closing","Délai Moy.","Marge (€)"]
for ci, h in enumerate(monthly_hdrs, 1):
    cell = ws_kpi.cell(row=r, column=ci, value=h)
    cell.font = Font(name="Arial", bold=True, size=9, color=C_WHITE)
    cell.fill = PatternFill("solid", fgColor=C_NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()
ws_kpi.row_dimensions[r].height = 20

for ri, row in monthly.iterrows():
    er = ri + 10
    vals = [row["month_label"], row["total_leads"], row["outbound_leads"],
            row["inbound_leads"], row["pct_inbound"], row["rdv"],
            row["signed"], row["conv_rate"], row["avg_delay"], row["margin"]]
    fmts = [None,"#,##0","#,##0","#,##0","0.0%","#,##0","#,##0","0.0%","0.0","#,##0 €"]
    bg_row = "F9FAFB" if ri % 2 == 0 else C_WHITE

    for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
        cell = ws_kpi.cell(row=er, column=ci, value=v)
        cell.font = Font(name="Arial", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
        if fmt: cell.number_format = fmt
        # Highlight inbound % with colour gradient
        if ci == 5:
            intensity = int(min(v * 255 * 3, 255))
            hex_g = f"{'%02X' % (255-intensity//2)}{'%02X' % (255-intensity//3)}{'%02X' % (255-intensity//4)}"
            cell.fill = PatternFill("solid", fgColor="D1FAE5" if v > 0.15 else "FFF9C4" if v > 0.05 else "FEE2E2")
        else:
            cell.fill = PatternFill("solid", fgColor=bg_row)
    ws_kpi.row_dimensions[er].height = 18

# Totals row
er = len(monthly) + 10
tot_vals = ["TOTAL", monthly["total_leads"].sum(), monthly["outbound_leads"].sum(),
            monthly["inbound_leads"].sum(), monthly["inbound_leads"].sum()/monthly["total_leads"].sum(),
            monthly["rdv"].sum(), monthly["signed"].sum(),
            monthly["signed"].sum()/monthly["total_leads"].sum(),
            monthly["avg_delay"].mean(), monthly["margin"].sum()]
tot_fmts = [None,"#,##0","#,##0","#,##0","0.0%","#,##0","#,##0","0.0%","0.0","#,##0 €"]
for ci, (v, fmt) in enumerate(zip(tot_vals, tot_fmts), 1):
    cell = ws_kpi.cell(row=er, column=ci, value=v)
    cell.font = Font(name="Arial", bold=True, size=9, color=C_WHITE)
    cell.fill = PatternFill("solid", fgColor=C_NAVY_L)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()
    if fmt: cell.number_format = fmt
ws_kpi.row_dimensions[er].height = 20

for i, w in enumerate([9,10,10,10,9,8,8,9,9,11], 1):
    ws_kpi.column_dimensions[get_column_letter(i)].width = w

# ── SHEET 3: SOURCE ANALYSIS ─────────────────────────────────────────────────
ws_src = wb.create_sheet("🎯 Analyse par Source")
ws_src.sheet_view.showGridLines = False

ws_src.merge_cells("A1:L1")
c = ws_src["A1"]
c.value = "VROOM MARKET  ·  Performance par Canal d'Acquisition  ·  2025"
c.font = Font(name="Arial", bold=True, size=12, color=C_WHITE)
c.fill = PatternFill("solid", fgColor=C_NAVY)
c.alignment = Alignment(horizontal="left", vertical="center")
ws_src.row_dimensions[1].height = 30

# Source performance table
src_perf = df.groupby("source").agg(
    total=("lead_id","count"),
    qualified=("qualified","sum"),
    rdv=("rdv","sum"),
    signed=("closed","sum"),
    avg_delay=("delay_days","mean"),
    inbound=("inbound","first"),
    cpl=("cpl_eur","first"),
    margin=("margin_eur","sum"),
).reset_index()
src_perf["taux_qual"] = src_perf["qualified"] / src_perf["total"]
src_perf["taux_rdv"]  = src_perf["rdv"] / src_perf["total"]
src_perf["taux_closing"] = src_perf["signed"] / src_perf["total"]
src_perf["type"] = src_perf["inbound"].map({True:"🟢 Inbound", False:"🔴 Outbound"})
src_perf = src_perf.sort_values("taux_closing", ascending=False)

hdrs_src = ["Source","Type","Total Leads","Qualifiés","RDV","Signés",
            "Taux Qualif.","Taux RDV","Taux Closing","Délai Moy. (j)","CPL (€)","Marge (€)"]
r = 3
ws_src.row_dimensions[r].height = 22
for ci, h in enumerate(hdrs_src, 1):
    cell = ws_src.cell(row=r, column=ci, value=h)
    cell.font = Font(name="Arial", bold=True, size=9, color=C_WHITE)
    cell.fill = PatternFill("solid", fgColor=C_NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()

PERF_COLORS = {
    "LeBonCoin":   ("FFF9C4","7A6200"),
    "LaCentrale":  ("FFF3E0","7C4A00"),
    "GBP":         ("D1FAE5","065F46"),
    "Facebook":    ("EDE9FE","4C1D95"),
    "Formulaire":  ("DBEAFE","1E3A8A"),
    "Partenariat": ("ECFDF5","052E16"),
}

for ri, row in src_perf.iterrows():
    er = ri + 4
    vals = [row["source"], row["type"], row["total"], row["qualified"],
            row["rdv"], row["signed"], row["taux_qual"], row["taux_rdv"],
            row["taux_closing"], row["avg_delay"], row["cpl"], row["margin"]]
    fmts = [None,None,"#,##0","#,##0","#,##0","#,##0","0.0%","0.0%","0.0%","0.0","#,##0 €","#,##0 €"]

    bg, fc = PERF_COLORS.get(row["source"], ("F9FAFB","000000"))

    for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
        cell = ws_src.cell(row=er, column=ci, value=v)
        cell.font = Font(name="Arial", size=10, bold=(ci in (1,9)), color=fc if ci==1 else "000000")
        cell.fill = PatternFill("solid", fgColor=bg if ci == 1 else "FFFFFF" if ri % 2 else "F9FAFB")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
        if fmt: cell.number_format = fmt

        # Colour-code taux closing
        if ci == 9:
            v_num = row["taux_closing"]
            if v_num >= 0.22:   cell.fill = PatternFill("solid", fgColor="D1FAE5"); cell.font = Font(name="Arial",bold=True,size=10,color="065F46")
            elif v_num >= 0.13: cell.fill = PatternFill("solid", fgColor="FEF3C7"); cell.font = Font(name="Arial",bold=True,size=10,color="7C4A00")
            else:               cell.fill = PatternFill("solid", fgColor="FEE2E2"); cell.font = Font(name="Arial",bold=True,size=10,color="991B1B")

    ws_src.row_dimensions[er].height = 22

widths_src = [13,14,11,10,8,8,11,10,11,13,9,11]
for i, w in enumerate(widths_src, 1):
    ws_src.column_dimensions[get_column_letter(i)].width = w

# ── SUMMARY BOXES ──
r_sum = 12
ws_src.merge_cells(f"A{r_sum}:D{r_sum}")
c = ws_src[f"A{r_sum}"]
c.value = "💡 INSIGHT CLÉS"
c.font = Font(name="Arial", bold=True, size=10, color=C_WHITE)
c.fill = PatternFill("solid", fgColor=C_NAVY_L)
c.alignment = Alignment(horizontal="left", vertical="center")
ws_src.row_dimensions[r_sum].height = 22

insights = [
    ("Meilleur canal", "Partenariat — 30% closing"),
    ("Inbound vs Outbound", "Inbound convertit 2× mieux"),
    ("CPL le plus bas", "GBP, Facebook, Partenariat → 0 €"),
    ("Canal à fort potentiel", "GBP : montée en puissance Q4"),
]
for i, (k, v) in enumerate(insights, 1):
    cell_k = ws_src.cell(row=r_sum+i, column=1, value=k)
    cell_k.font = Font(name="Arial", bold=True, size=9, color="374151")
    cell_k.fill = PatternFill("solid", fgColor="F3F4F6")
    cell_k.alignment = Alignment(horizontal="left", vertical="center")
    cell_k.border = thin_border()
    ws_src.merge_cells(f"A{r_sum+i}:B{r_sum+i}")

    cell_v = ws_src.cell(row=r_sum+i, column=3, value=v)
    cell_v.font = Font(name="Arial", size=9, color=C_INBOUND)
    cell_v.fill = PatternFill("solid", fgColor="ECFDF5")
    cell_v.alignment = Alignment(horizontal="left", vertical="center")
    cell_v.border = thin_border()
    ws_src.merge_cells(f"C{r_sum+i}:F{r_sum+i}")
    ws_src.row_dimensions[r_sum+i].height = 20

# ── SHEET 4: FUNNEL ────────────────────────────────────────────────────────
ws_fun = wb.create_sheet("🔻 Funnel Conversion")
ws_fun.sheet_view.showGridLines = False

ws_fun.merge_cells("A1:J1")
c = ws_fun["A1"]
c.value = "VROOM MARKET  ·  Analyse du Funnel de Conversion  ·  2025"
c.font = Font(name="Arial", bold=True, size=12, color=C_WHITE)
c.fill = PatternFill("solid", fgColor=C_NAVY)
c.alignment = Alignment(horizontal="left", vertical="center")
ws_fun.row_dimensions[1].height = 30

# Global funnel
funnel_stages = ["Leads reçus","Contactés","Qualifiés","RDV planifié","RDV effectué","Signés"]
total_c = len(df)
contacted = df[df["status"] != "NRP"].shape[0]
qualified_n = df["qualified"].sum()
rdv_planned = df[df["status"].isin(["RDV planifié","RDV effectué","Signé"])].shape[0]
rdv_done = df[df["status"].isin(["RDV effectué","Signé"])].shape[0]
signed_n = df["closed"].sum()

funnel_vals = [total_c, contacted, qualified_n, rdv_planned, rdv_done, signed_n]
funnel_pcts = [1.0] + [v/total_c for v in funnel_vals[1:]]
funnel_drop = [None] + [(funnel_vals[i-1]-v)/funnel_vals[i-1] for i,v in enumerate(funnel_vals[1:],1)]

ws_fun.row_dimensions[3].height = 22
hdrs_fun = ["Étape","Leads","% du Total","Conversion vs étape préc.","Perte vs étape préc."]
for ci, h in enumerate(hdrs_fun, 1):
    cell = ws_fun.cell(row=3, column=ci, value=h)
    cell.font = Font(name="Arial", bold=True, size=9, color=C_WHITE)
    cell.fill = PatternFill("solid", fgColor=C_NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()

FUNNEL_BG = ["1E3A8A","1D4ED8","2563EB","3B82F6","60A5FA","93C5FD"]
FUNNEL_FC = [C_WHITE,C_WHITE,C_WHITE,C_WHITE,"1E3A8A","1E3A8A"]

for i, (stage, val, pct) in enumerate(zip(funnel_stages, funnel_vals, funnel_pcts)):
    er = i + 4
    drop = funnel_drop[i]
    step_conv = None if i == 0 else funnel_vals[i]/funnel_vals[i-1]

    row_data = [stage, val, pct, step_conv, drop]
    fmts     = [None, "#,##0", "0.0%", "0.0%", "0.0%"]

    for ci, (v, fmt) in enumerate(zip(row_data, fmts), 1):
        cell = ws_fun.cell(row=er, column=ci, value=v if v is not None else "—")
        cell.font = Font(name="Arial", size=10, bold=(ci==1), color=C_WHITE if i < 4 else "1E3A8A")
        fill_c = FUNNEL_BG[i] if ci == 1 else ("F9FAFB" if i % 2 else "FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill_c)
        cell.alignment = Alignment(horizontal="center" if ci>1 else "left", vertical="center")
        cell.border = thin_border()
        if fmt and v is not None and v != "—": cell.number_format = fmt
    ws_fun.row_dimensions[er].height = 22

# Funnel by source comparison
r_fs = 12
ws_fun.merge_cells(f"A{r_fs}:J{r_fs}")
c = ws_fun[f"A{r_fs}"]
c.value = "COMPARAISON DU FUNNEL : INBOUND vs OUTBOUND"
c.font = Font(name="Arial", bold=True, size=10, color=C_WHITE)
c.fill = PatternFill("solid", fgColor=C_NAVY_L)
c.alignment = Alignment(horizontal="left", vertical="center")
ws_fun.row_dimensions[r_fs].height = 22

for col_type, (label, is_in, col_offset) in enumerate([("OUTBOUND",False,1),("INBOUND",True,4)]):
    sub = df[df["inbound"]==is_in]
    sc = contacted if not is_in else df[(df["inbound"]==True) & (df["status"]!="NRP")].shape[0]
    sq = sub["qualified"].sum()
    sr = sub["rdv"].sum()
    ss = sub["closed"].sum()
    total_sub = len(sub)

    r_h = r_fs + 1
    ws_fun.cell(row=r_h, column=col_offset, value=label).font = Font(name="Arial",bold=True,size=10,color=C_WHITE if is_in else "000000")
    ws_fun.cell(row=r_h, column=col_offset).fill = PatternFill("solid", fgColor=C_INBOUND if is_in else C_OUTBOUND)
    ws_fun.cell(row=r_h, column=col_offset).alignment = Alignment(horizontal="center",vertical="center")
    ws_fun.row_dimensions[r_h].height = 20

    stages_sub = [("Total leads",total_sub,1.0),("Contactés",sc,sc/total_sub),
                  ("Qualifiés",sq,sq/total_sub),("RDV",sr,sr/total_sub),("Signés",ss,ss/total_sub)]
    for ri2, (s2, v2, p2) in enumerate(stages_sub):
        er2 = r_h + 1 + ri2
        ws_fun.cell(row=er2, column=col_offset, value=s2).font = Font(name="Arial",size=9)
        ws_fun.cell(row=er2, column=col_offset).border = thin_border()
        ws_fun.cell(row=er2, column=col_offset).alignment = Alignment(horizontal="left",vertical="center")

        vc = ws_fun.cell(row=er2, column=col_offset+1, value=v2)
        vc.number_format = "#,##0"; vc.border = thin_border()
        vc.alignment = Alignment(horizontal="center",vertical="center")

        pc = ws_fun.cell(row=er2, column=col_offset+2, value=p2)
        pc.number_format = "0.0%"; pc.border = thin_border()
        pc.alignment = Alignment(horizontal="center",vertical="center")
        if p2 > 0.15: pc.fill = PatternFill("solid",fgColor="D1FAE5")
        ws_fun.row_dimensions[er2].height = 18

for i,w in enumerate([18,10,8,18,10,8,18,10,8,10],1):
    ws_fun.column_dimensions[get_column_letter(i)].width = w

# ── SHEET 5: ROI ANALYSIS ──────────────────────────────────────────────────
ws_roi = wb.create_sheet("💰 ROI & Coût par Lead")
ws_roi.sheet_view.showGridLines = False

ws_roi.merge_cells("A1:H1")
c = ws_roi["A1"]
c.value = "VROOM MARKET  ·  Analyse ROI & Coût par Acquisition  ·  2025"
c.font = Font(name="Arial", bold=True, size=12, color=C_WHITE)
c.fill = PatternFill("solid", fgColor=C_NAVY)
c.alignment = Alignment(horizontal="left", vertical="center")
ws_roi.row_dimensions[1].height = 30

# CPL / CPA table
roi_data = []
for src in SOURCE_ORDER:
    s = df[df["source"]==src]
    n = len(s)
    closed_n = s["closed"].sum()
    cpl = s["cpl_eur"].iloc[0]
    total_cost = n * cpl
    cpa = total_cost / closed_n if closed_n > 0 else 0
    marge = s["margin_eur"].sum()
    roi_val = (marge - total_cost) / total_cost if total_cost > 0 else float('inf')
    roi_data.append({
        "Source": src,
        "Type": "Inbound" if s["inbound"].iloc[0] else "Outbound",
        "Leads": n,
        "Signés": closed_n,
        "CPL (€)": cpl,
        "Coût total (€)": total_cost,
        "CPA (€)": round(cpa),
        "Marge brute (€)": round(marge) if not pd.isna(marge) else 0,
        "ROI estimé": roi_val if total_cost > 0 else "∞ (gratuit)",
    })

r = 3
hdrs_roi = ["Source","Type","Leads","Signés","CPL (€)","Coût total (€)","CPA (€)","Marge brute (€)","ROI estimé"]
ws_roi.row_dimensions[r].height = 22
for ci, h in enumerate(hdrs_roi, 1):
    cell = ws_roi.cell(row=r, column=ci, value=h)
    cell.font = Font(name="Arial", bold=True, size=9, color=C_WHITE)
    cell.fill = PatternFill("solid", fgColor=C_NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()

for ri, row_r in enumerate(roi_data):
    er = ri + 4
    is_in = row_r["Type"] == "Inbound"
    bg_r = "ECFDF5" if is_in else ("F9FAFB" if ri%2 else "FFFFFF")

    for ci, key in enumerate(hdrs_roi, 1):
        cell = ws_roi.cell(row=er, column=ci)
        v = row_r[key]
        fmts = {
            "Leads":"#,##0","Signés":"#,##0","CPL (€)":"#,##0 €",
            "Coût total (€)":"#,##0 €","CPA (€)":"#,##0 €","Marge brute (€)":"#,##0 €"
        }
        if key == "ROI estimé":
            if isinstance(v, str):
                cell.value = v
                cell.font = Font(name="Arial", size=10, bold=True, color="065F46")
                cell.fill = PatternFill("solid", fgColor="D1FAE5")
            else:
                cell.value = v
                cell.number_format = "0.0×" if v != 0 else "0"
                if v > 5:   cell.fill = PatternFill("solid",fgColor="D1FAE5"); cell.font = Font(name="Arial",size=10,bold=True,color="065F46")
                elif v > 1: cell.fill = PatternFill("solid",fgColor="FEF3C7"); cell.font = Font(name="Arial",size=10,color="7C4A00")
                else:        cell.fill = PatternFill("solid",fgColor="FEE2E2"); cell.font = Font(name="Arial",size=10,color="991B1B")
        else:
            cell.value = v
            if key in fmts and isinstance(v, (int,float)):
                cell.number_format = fmts[key]
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", fgColor=bg_r)

        cell.alignment = Alignment(horizontal="center" if ci>1 else "left", vertical="center")
        cell.border = thin_border()
    ws_roi.row_dimensions[er].height = 22

# ROI note
r_note = len(roi_data) + 6
ws_roi.merge_cells(f"A{r_note}:H{r_note}")
note_cell = ws_roi[f"A{r_note}"]
note_cell.value = "⚠️  CPL (Coût Par Lead) outbound estimé à 8 € = valorisation du temps de prospection (20 min × 24 €/h).  CPA = Coût total ÷ Deals signés.  ROI = (Marge brute − Coût acquisition) ÷ Coût acquisition.  Marge brute estimée 8–14% de la valeur véhicule."
note_cell.font = Font(name="Arial", size=8, italic=True, color="6B7280")
note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_roi.row_dimensions[r_note].height = 30

for i,w in enumerate([13,10,8,8,9,12,10,13,11],1):
    ws_roi.column_dimensions[get_column_letter(i)].width = w

# ── SHEET 6: INBOUND EVOLUTION ────────────────────────────────────────────
ws_evol = wb.create_sheet("📈 Évolution Inbound")
ws_evol.sheet_view.showGridLines = False

ws_evol.merge_cells("A1:N1")
c = ws_evol["A1"]
c.value = "VROOM MARKET  ·  Montée en Puissance des Canaux Inbound  ·  Jan–Déc 2025"
c.font = Font(name="Arial", bold=True, size=12, color=C_WHITE)
c.fill = PatternFill("solid", fgColor=C_NAVY)
c.alignment = Alignment(horizontal="left", vertical="center")
ws_evol.row_dimensions[1].height = 30

# Monthly by source
pivot = df.pivot_table(index="month_num", columns="source", values="lead_id",
                       aggfunc="count", fill_value=0).reset_index()
inbound_sources = ["GBP","Facebook","Formulaire","Partenariat"]
outbound_sources = ["LeBonCoin","LaCentrale"]

r = 3
ws_evol.row_dimensions[r].height = 22

hdrs_e = ["Mois"] + SOURCE_ORDER + ["Total Inbound","Total Outbound","% Inbound"]
for ci, h in enumerate(hdrs_e, 1):
    cell = ws_evol.cell(row=r, column=ci, value=h)
    cell.font = Font(name="Arial", bold=True, size=9, color=C_WHITE)
    is_in_src = h in inbound_sources
    is_out_src = h in outbound_sources
    bg_h = C_INBOUND if is_in_src else (C_OUTBOUND if is_out_src else C_NAVY)
    cell.fill = PatternFill("solid", fgColor=bg_h)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()

for ri, row_p in pivot.iterrows():
    er = ri + 4
    m_label = MONTH_LABELS[int(row_p["month_num"])-1]
    row_vals = [m_label]
    for src in SOURCE_ORDER:
        row_vals.append(int(row_p.get(src, 0)))

    in_total = sum(int(row_p.get(s,0)) for s in inbound_sources)
    out_total = sum(int(row_p.get(s,0)) for s in outbound_sources)
    pct = in_total / (in_total + out_total) if (in_total+out_total) > 0 else 0
    row_vals += [in_total, out_total, pct]
    fmts_e = [None]+["#,##0"]*6+["#,##0","#,##0","0.0%"]

    for ci, (v, fmt) in enumerate(zip(row_vals, fmts_e), 1):
        cell = ws_evol.cell(row=er, column=ci, value=v)
        src_name = SOURCE_ORDER[ci-2] if 2 <= ci <= 7 else None
        is_in_src = src_name in inbound_sources if src_name else (ci == 8)
        bg = "ECFDF5" if is_in_src else ("FFF9C4" if ci <= 7 else "F9FAFB")
        if ci == 10:  # % inbound column
            bg = "D1FAE5" if v > 0.15 else "FEF3C7" if v > 0.05 else "FEE2E2"
        cell.fill = PatternFill("solid", fgColor=bg if ri%2==0 else "FFFFFF")
        cell.font = Font(name="Arial", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
        if fmt: cell.number_format = fmt
    ws_evol.row_dimensions[er].height = 18

for i,w in enumerate([7,11,11,8,9,11,12,12,12,10],1):
    ws_evol.column_dimensions[get_column_letter(i)].width = w

# ── SHEET 7: STORYTELLING / README ────────────────────────────────────────
ws_story = wb.create_sheet("📖 Contexte & Méthodologie")
ws_story.sheet_view.showGridLines = False

def story_block(ws, r, c, text, bg, fc, bold=False, sz=10, merge_to=None, height=22, wrap=False, italic=False, halign="left"):
    if merge_to:
        ws.merge_cells(f"{get_column_letter(c)}{r}:{get_column_letter(merge_to)}{r}")
    cell = ws.cell(row=r, column=c, value=text)
    cell.font = Font(name="Arial", bold=bold, italic=italic, size=sz, color=fc)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    ws.row_dimensions[r].height = height
    return cell

story_block(ws_story, 1, 1, "VROOM MARKET  ·  Contexte du Projet & Méthodologie Data", C_NAVY, C_WHITE, True, 13, merge_to=10, height=34)

sections = [
    (3, "🏢 CONTEXTE BUSINESS", C_NAVY_L, C_WHITE, True, 11, 26),
    (4, "Vroom Market est une PME spécialisée dans le rachat de véhicules d'occasion auprès de particuliers.", C_BG, "374151", False, 10, 20),
    (5, "Situation initiale (Jan 2025) : 100% des leads proviennent de plateformes outbound (Le Bon Coin, La Centrale) via scraping CRM + extension Chrome.", C_BG, "374151", False, 10, 36),
    (6, "Problème : forte dépendance aux marketplaces tiers, coût d'acquisition caché (temps humain non mesuré), 0% de leads inbound.", C_BG, "991B1B", False, 10, 36),

    (8, "📊 STRATÉGIE DATA MISE EN PLACE", C_NAVY_L, C_WHITE, True, 11, 26),
    (9, "Objectif : Mettre en place des canaux inbound low-cost et mesurer leur montée en puissance vs l'outbound existant.", "ECFDF5", "065F46", False, 10, 28),
    (10, "Canaux activés (progressivement) : Google Business Profile (SEO local) · Facebook Marketplace · Formulaire web (landing page) · Partenariats locaux (garages, CT)", C_BG, "374151", False, 10, 36),
    (11, "Tracking : tag source obligatoire sur chaque lead dans le CRM BWA dès le 1er janvier 2025.", C_BG, "374151", False, 10, 20),

    (13, "🔢 MÉTHODOLOGIE DE SIMULATION DES DONNÉES", C_NAVY_L, C_WHITE, True, 11, 26),
    (14, "Le dataset (1 019 leads sur 12 mois) est simulé de manière réaliste à partir de benchmarks sectoriels :", C_BG, "374151", False, 10, 20),
    (15, "  · Taux de conversion inbound vs outbound : basés sur études HubSpot, Salesforce & benchmarks lead gen automobile France", C_BG, "374151", False, 9, 20),
    (16, "  · CPL outbound (8 €) : valorisation du temps de prospection (20 min × 24 €/h salaire chargé)", C_BG, "374151", False, 9, 20),
    (17, "  · Valeur véhicule (3 500–12 000 €) et marge (6–14%) : cohérents avec le marché VO local France 2025", C_BG, "374151", False, 9, 20),
    (18, "  · Ramp-up inbound : démarrage quasi nul (Jan–Feb), montée progressive, +80% vs base en Dec — reflète la réalité d'un canal SEO/social organique", C_BG, "374151", False, 9, 20),

    (20, "📈 RÉSULTATS OBSERVÉS (simulés)", C_NAVY_L, C_WHITE, True, 11, 26),
    (21, "✓ Part inbound passe de ~2% en Jan à ~22% en Déc après activation des canaux", "D1FAE5", "065F46", False, 10, 20),
    (22, "✓ Taux de closing inbound : ~24% vs ~14% outbound — soit un avantage de conversion ×1.7", "D1FAE5", "065F46", False, 10, 20),
    (23, "✓ ROI inbound : ∞ sur canaux gratuits (GBP, Facebook, Partenariats) vs ROI outbound estimé à ~×6", "D1FAE5", "065F46", False, 10, 20),
    (24, "✓ Délai moyen de traitement inbound : 4.2j vs 6.8j outbound (leads plus engagés dès le départ)", "D1FAE5", "065F46", False, 10, 20),

    (26, "🛠️ OUTILS & STACK", C_NAVY_L, C_WHITE, True, 11, 26),
    (27, "Python (pandas, numpy) · openpyxl · Looker Studio / Power BI pour dashboards · CRM BWA (fictif) · Git + GitHub", C_BG, "374151", False, 10, 20),

    (29, "⚠️  AVERTISSEMENT", "FEF3C7", "7C4A00", True, 10, 22),
    (30, "Ce projet est basé sur des données entièrement simulées à des fins de démonstration portfolio. Les noms d'entreprises, volumes et chiffres sont fictifs mais construits sur des benchmarks réels du marché.", "FEF9C3", "7C4A00", False, 9, 40),
]

for item in sections:
    r_s, txt, bg_s, fc_s, bold_s, sz_s, h_s = item
    story_block(ws_story, r_s, 1, txt, bg_s, fc_s, bold_s, sz_s, merge_to=10, height=h_s, wrap=True)

ws_story.column_dimensions["A"].width = 100

# ── SAVE ──────────────────────────────────────────────────────────────────────
out_path = "/home/claude/vroom-portfolio/dashboards/vroom_market_analytics_2025.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
